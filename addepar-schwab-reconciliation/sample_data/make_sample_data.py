#!/usr/bin/env python3
"""
Generate a small, fully synthetic sample dataset for the reconciliation demo.

Nothing here is real: the securities are well-known public instruments, and the
internal identifiers (Addepar Entity IDs, Schwab Security Numbers) are made up.
The rows are chosen to trigger every interesting branch of reconcile.py:

  * a CUSIP mangled into scientific notation (repaired from the US ISIN)
  * an Addepar row missing its CUSIP (backfilled from Schwab)
  * an Addepar row missing its ISIN (backfilled from Schwab)
  * a security that only overlaps on ticker (low-confidence match)
  * a matched security with a ticker conflict (Addepar carries a stale ticker)
  * duplicate Schwab rows for one security (collapsed to the modal value)
  * an unmatched security with good identifiers (held-away / not custodied)
  * an unmatched security with only a ticker (not found at the custodian)
  * a Schwab Security Number supplied by hand in the Addepar export (overlay)

Run:  python make_sample_data.py
"""

from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
DATE = "20260115"


# --- Schwab flat-file writer (pipe-delimited, two-row header) ---------------
def write_schwab_file(path: Path, rows):
    """rows: list of (description, ticker, cusip, isin, schwab_sec_nbr)."""
    lines = []
    lines.append("H1|Sample Custodial Reporting Extract")
    # H2 = column groups, H3 = column names; combined labels drive the parser
    lines.append("H2|Security|Security|||Schwab|")
    lines.append("H3|Description|Ticker Symbol|CUSIP|ISIN|Sec Nbr|Currency")
    for desc, tkr, cusip, isin, sec in rows:
        lines.append(f"DL|{desc}|{tkr}|{cusip}|{isin}|{sec}|USD")
    lines.append(f"T1|{len(rows)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {path.name} ({len(rows)} DL rows)")


# Taxable lots (ULT)
ult_rows = [
    ("APPLE INC",                 "AAPL", "037833100", "US0378331005", "1000001"),
    ("MICROSOFT CORP",            "MSFT", "594918104", "US5949181045", "1000002"),
    ("AMAZON COM INC",            "AMZN", "023135106", "US0231351067", "1000003"),
    ("VANGUARD S&P 500 ETF",      "VOO",  "922908363", "US9229083632", "1000004"),
    ("SPDR S&P 500 ETF TRUST",    "SPY",  "78462F103", "US78462F1030", "1000005"),
    ("NVIDIA CORP",               "NVDA", "67066G104", "US67066G1040", "1000006"),
    # duplicate of AAPL: exercises modal collapse in build_schwab_maps
    ("APPLE INC",                 "AAPL", "037833100", "US0378331005", "1000001"),
    ("TESLA INC",                 "TSLA", "88160R101", "US88160R1014", "1000007"),
    # ticker-only: no CUSIP / ISIN on the Schwab side
    ("BERKSHIRE HATHAWAY CL B",   "BRKB", "",          "",             "1000008"),
    ("ISHARES CORE US AGG BOND",  "AGG",  "464287226", "US4642872265", "1000009"),
    ("META PLATFORMS INC",        "META", "30303M102", "US30303M1027", "1000012"),
]

# Non-taxable lots (ULN)
uln_rows = [
    ("JOHNSON & JOHNSON",         "JNJ",  "478160104", "US4781601046", "1000010"),
    ("VANGUARD TOTAL STOCK ETF",  "VTI",  "922908769", "US9229087690", "1000011"),
    ("MICROSOFT CORP",            "MSFT", "594918104", "US5949181045", "1000002"),
]

write_schwab_file(HERE / f"CRS{DATE}.ULT", ult_rows)
write_schwab_file(HERE / f"CRS{DATE}.ULN", uln_rows)


# --- Addepar "Portfolio View" export ----------------------------------------
# Columns: Security | Entity ID | Schwab Security Number | Ticker | ISIN | CUSIP
addepar_rows = [
    # Security,                    EntityID,   Schwab#,  Ticker, ISIN,           CUSIP
    ("Apple Inc",                  "20000001", "",       "AAPL", "US0378331005", "3.7833E+08"),  # scientific -> repaired
    ("Microsoft Corp",             "20000002", "",       "MSFT", "US5949181045", "594918104"),
    ("Amazon.com Inc",             "20000003", "",       "AMZN", "US0231351067", ""),            # CUSIP backfilled
    ("Vanguard S&P 500 ETF",       "20000004", "",       "VOO",  "US9229083632", "922908363"),
    ("SPDR S&P 500 ETF Trust",     "20000005", "",       "SPY",  "",             "78462F103"),   # ISIN backfilled
    ("NVIDIA Corp",                "20000006", "",       "NVDA", "US67066G1040", "67066G104"),
    ("Tesla Inc",                  "20000007", "",       "TSLA", "US88160R1014", "88160R101"),
    ("Berkshire Hathaway Cl B",    "20000008", "",       "BRKB", "",             ""),             # ticker-only match
    ("iShares Core US Agg Bond",   "20000009", "",       "AGG",  "US4642872265", "464287226"),
    ("Johnson & Johnson",          "20000010", "",       "JNJ",  "US4781601046", "478160104"),
    ("Vanguard Total Stock ETF",   "20000011", "",       "VTI",  "US9229087690", "922908769"),
    ("Emerging Micro Cap Fund",    "20000012", "",       "ZZZZ", "",             ""),             # unmatched (ticker)
    ("Municipal Bond 2031",        "20000013", "9999999","MUNI1","",             ""),             # manual overlay
    # Addepar still carries the old ticker (FB); Schwab has the current one (META).
    # Matches on ISIN + CUSIP, but the ticker conflict is flagged.
    ("Meta Platforms Inc",         "20000014", "",       "FB",   "US30303M1027", "30303M102"),    # matched, conflict
    # Held at another custodian / not custodied at Schwab: good identifiers, no Schwab record.
    ("Global Macro Fund Ltd",      "20000015", "",       "GMFXX","IE00B4L5Y983", ""),             # unmatched (held-away)
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Portfolio View"
ws.append(["Security", "Entity ID", "Schwab Security Number", "Ticker Symbol", "ISIN", "CUSIP"])
for row in addepar_rows:
    # write every cell as text so leading zeros and the scientific string survive
    ws.append([str(v) for v in row])
# force text format on the CUSIP column so "3.7833E+08" is stored as a string
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=6).number_format = "@"

out_xlsx = HERE / "addepar_export.xlsx"
wb.save(out_xlsx)
print(f"wrote {out_xlsx.name} ({len(addepar_rows)} securities)")
